import unittest
from selenium import webdriver
import time
import openpyxl

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import logging
import HtmlTestRunner

class HitPageTest(unittest.TestCase):

    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome(executable_path="C:/Selenium Drivers/chromedriver")
        cls.driver.get("https://is.hit.ac.il/nidp/idff/sso?id=HITUserPassword&sid=0&option=credential&sid=0&target=https%3A%2F%2Fportal.hit.ac.il%2F")
        workbook = openpyxl.load_workbook("res/hit.xlsx")
        cls.hitSheet = workbook ["גיליון2"]
        logging.basicConfig(filename="logs/wallashops logging.log",
                            # הוספנו תיקייה שקראנו לה logs, מוסיפים תיקייה קליק ימני על הפרויקט my directory
                            filemode="w", format="%(asctime)s: %(levelname)s: %(message)s", level=logging.DEBUG, )
        cls.logger = logging.getLogger()
        cls.logger.debug("Finished initializing")

    @classmethod
    def tearDown(cls):
        cls.driver.quit()

    def test_fakeLogin (self):
        self.logger.debug("In test_EmptyPassword()")
        self.driver.find_element_by_id('username').send_keys("Name")
        self.driver.find_element_by_id('password').send_keys("passs")
        self.driver.find_element_by_xpath("//input[@value='כניסה - Login']").click()
        errorElement = self.driver.find_element_by_id("errorDIV")
        self.assertTrue("Invalid Credentials, please try again" in errorElement.text, "Wrong error message")

    def test_goodLogin (self):
        userName = self.hitSheet.cell(2, 1).value
        password = self.hitSheet.cell(2, 2).value
        self.driver.find_element_by_id('username').send_keys(userName)
        self.driver.find_element_by_id('password').send_keys(password)
        self.driver.find_element_by_xpath("//input[@value='כניסה - Login']").click()
        time.sleep(3)
        self.assertEqual(self.driver.title, "", "Not main page")

    def goodLoginByid(self, idText):
        try:
            self.driver.find_element_by_id(idText)
            return True
        except NoSuchElementException:
            return False

    def goodLoginByXpath(self, Xpath):
        try:
            self.driver.find_element_by_xpath(Xpath)
            return True
        except NoSuchElementException:
            return False

    def test_MouseOver (self):
        self.test_goodLogin()
        actions = ActionChains(self.driver)
        menuElement = self.driver.find_element_by_id("HIT_settings")
        actions.move_to_element(menuElement).perform()
        myFilesElement = self.driver.find_element_by_xpath("//div[text()='הקבצים שלי']")
        self.assertTrue(myFilesElement.is_displayed(), "Bug - Popup not shown when mouse is over setting icon")

    def test_goodLoginByFuncation (self):
        self.driver.find_element_by_id('username').send_keys("302885041")
        self.driver.find_element_by_id('password').send_keys("sahar16")
        self.driver.find_element_by_xpath("//input[@value='כניסה - Login']").click()
        self.driver.find_element_by_xpath('//div[text() = "התנתק"]')
        time.sleep(3)
        self.assertTrue(self.goodLoginByXpath, "Not main page")

    def goodLoginByid(self, idText): # עושים try וexcept כדי, שאם התוכנית לא תמצא את האלמנט היא לא תעצור ותקרוס, אלא תמשיך לשאר שורות הקוד.
        try:
            self.driver.find_element_by_id(idText)
            return True
        except NoSuchElementException:
            return False

    def goodLoginByXpath(self, Xpath):
        try:
            self.driver.find_element_by_xpath(Xpath)
            return True
        except NoSuchElementException:
            return False

if __name__ == "__main__":
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output="..\\logs"))
