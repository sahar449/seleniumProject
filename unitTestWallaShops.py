import unittest
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
import logging

class WallaShopsTestPage (unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.driver = webdriver.Chrome(executable_path="C:/Selenium Drivers/chromedriver")
        cls.driver.get("https://www.wallashops.co.il/Login/")
        workbook = openpyxl.load_workbook("res/users.xlsx")
        cls.wallaSheet = workbook["Wallashops"]
        logging.basicConfig(filename="logs/wallashops logging.log",   #הוספנו תיקייה שקראנו לה logs, מוסיפים תיקייה קליק ימני על הפרויקט my directory
                filemode="w", format="%(asctime)s: %(levelname)s: %(message)s", level=logging.DEBUG, )
        cls.logger = logging.getLogger()
        cls.logger.debug("Finished initializing")


    @classmethod
    def tearDownClass(cls):
        cls.driver.quit()

    # @unittest.SkipTest  # לדלג על הטסט
    def test_EmptyUser(self):
        self.driver.find_element_by_id('txtUserName').send_keys("")
        self.driver.find_element_by_id('btnSubmit').click()
        errorData = self.driver.find_element_by_xpath('//span [text() = "יש להזין תעודת זהות או אימייל"]')
        self.assertTrue("יש להזין תעודת זהות או אימייל" in errorData.text, "Wrong error message")

    def test_EmptyPassword(self):
        self.logger.debug("In test_EmptyPassword()")
        self.driver.find_element_by_id('txtPassword').send_keys('')
        self.driver.find_element_by_id('btnSubmit').click()
        errorData = self.driver.find_element_by_xpath('//span [text() = "אי אפשר להכנס בלי סיסמה..."]')
        self.assertTrue("אי אפשר להכנס בלי סיסמה..." in errorData.text, "Wrong error message")

    def test_EmptyUserAndPassword(self):
        self.driver.find_element_by_id('txtUserName').send_keys("")
        self.driver.find_element_by_id('txtPassword').send_keys("")
        self.driver.find_element_by_id('btnSubmit').click()
        errorUser = self.driver.find_element_by_xpath('//span [text() = "יש להזין תעודת זהות או אימייל"]')
        self.assertTrue("יש להזין תעודת זהות או אימייל" in errorUser.text, "Wrong error message")
        errorPass = self.driver.find_element_by_xpath('//span [text() = "אי אפשר להכנס בלי סיסמה..."]')
        self.assertTrue("אי אפשר להכנס בלי סיסמה..." in errorPass.text, "Wrong error message")

    def test_GoodLogin(self):
        userName = self.wallaSheet.cell(2, 1).value
        password = self.wallaSheet.cell(2, 2).value
        self.driver.find_element_by_id('txtUserName').send_keys(userName)
        self.driver.find_element_by_id('txtPassword').send_keys(password)
        self.driver.find_element_by_id('btnSubmit').click()
        time.sleep(3)
        self.assertTrue(self.isElementExistByXpath('//a[@class="disconnectReq"]'))

    def isElementExistsById(self, idText):
        try:
            self.driver.find_element_by_id(idText)
            return True
        except NoSuchElementException: # כדי למצוא את שם השגיאה צריך ללחוץ על שם האלמנט נניח id וכפתור שמאלי בעכבר ובחלון החדש :Raises:
            return False

    def isElementExistByXpath(self, xpath):
        try:
            self.driver.find_element_by_xpath(xpath)
            return True
        except NoSuchElementException:
            return False

    def test_MouseOver(self):
        self.test_GoodLogin()
        actions = ActionChains(self.driver)
        menuElement = self.driver.find_element_by_xpath('//li [@menu-item-id= "4131"]')
        actions.move_to_element(menuElement).perform()
        NutritionElement = self.driver.find_element_by_xpath('//a [@href="/דיל-יומי/תוספי-תזונה-וויטמינים"]')
        self.assertTrue(NutritionElement.is_displayed(), "Bug - pop doesn't exist")

    def test_Asserations(self):
        price = 100
        self.assertGreaterEqual(price, 150, "price is lower than 150")


if __name__ == "__main__":
    unittest.main()